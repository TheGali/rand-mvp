"""Excel output writers for command center and funding summary.

Produces two workbooks:
- command_center.xlsx: one row per observation with all structured fields
- funding_summary.xlsx: costs grouped by system and priority timeframe
"""

from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers


def write_command_center(report, output_path):
    """Write the command center spreadsheet — one row per processed observation.

    Args:
        report: Report with populated processed_observations
        output_path: Path for the output .xlsx
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Command Center"

    # Headers
    headers = [
        "Obs #", "System", "Component", "Location", "Condition",
        "Observation Summary", "Recommendation", "Priority",
        "Cost Low", "Cost High", "Photo Ref"
    ]

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

    # Data rows
    for row_idx, po in enumerate(report.processed_observations, 2):
        # Show cost cells as empty when cost is 0 (no engineer-stated cost)
        cost_low_display = po.cost_low if po.cost_low else ""
        cost_high_display = po.cost_high if po.cost_high else ""

        values = [
            po.obs_number,
            po.system,
            po.component,
            po.location,
            po.condition,
            po.prose[:200] + "..." if len(po.prose) > 200 else po.prose,
            po.recommendation,
            po.priority,
            cost_low_display,
            cost_high_display,
            f"Photo {po.obs_number}",
        ]

        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical="top")

        # Currency formatting for cost columns (only when there's a value)
        if cost_low_display != "":
            ws.cell(row=row_idx, column=9).number_format = '$#,##0'
        if cost_high_display != "":
            ws.cell(row=row_idx, column=10).number_format = '$#,##0'

    # Auto-width columns (approximate)
    col_widths = [8, 22, 22, 25, 12, 50, 40, 15, 14, 14, 12]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    print(f"  Written: {output_path}")


def write_funding_summary(report, output_path):
    """Write the funding summary spreadsheet — costs grouped by system and timeframe.

    Timeframe columns are dynamic — derived from whatever the AI returned.

    Args:
        report: Report with populated processed_observations
        output_path: Path for the output .xlsx
    """
    # Collect all unique timeframes and group data
    # Skip observations with 0 cost (no engineer-stated cost) so totals aren't polluted
    system_timeframe_costs = defaultdict(lambda: defaultdict(float))
    all_timeframes = set()

    for po in report.processed_observations:
        timeframe = po.priority if po.priority else "Unclassified"
        system = po.system if po.system else "Other"
        all_timeframes.add(timeframe)
        if po.cost_high:
            system_timeframe_costs[system][timeframe] += po.cost_high

    # Sort timeframes — try to put them in chronological order
    def timeframe_sort_key(tf):
        tf_lower = tf.lower()
        if "immediate" in tf_lower or "1-2" in tf_lower or "year 1" in tf_lower:
            return 0
        if "1-3" in tf_lower or "short" in tf_lower:
            return 1
        if "3-5" in tf_lower or "medium" in tf_lower:
            return 2
        if "6-10" in tf_lower or "long" in tf_lower:
            return 3
        if "10" in tf_lower or "capital" in tf_lower:
            return 4
        return 5

    sorted_timeframes = sorted(all_timeframes, key=timeframe_sort_key)

    wb = Workbook()
    ws = wb.active
    ws.title = "Funding Summary"

    # Headers
    headers = ["System"] + sorted_timeframes + ["Total"]

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

    # Data rows — one per system
    sorted_systems = sorted(system_timeframe_costs.keys())
    for row_idx, system in enumerate(sorted_systems, 2):
        ws.cell(row=row_idx, column=1, value=system).border = thin_border

        row_total = 0
        for tf_idx, tf in enumerate(sorted_timeframes, 2):
            cost = system_timeframe_costs[system].get(tf, 0)
            cell = ws.cell(row=row_idx, column=tf_idx, value=cost)
            cell.number_format = '$#,##0'
            cell.border = thin_border
            row_total += cost

        total_cell = ws.cell(row=row_idx, column=len(headers), value=row_total)
        total_cell.number_format = '$#,##0'
        total_cell.font = Font(bold=True)
        total_cell.border = thin_border

    # Grand total row
    total_row = len(sorted_systems) + 2
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=total_row, column=1).border = thin_border

    for tf_idx, tf in enumerate(sorted_timeframes, 2):
        col_total = sum(
            system_timeframe_costs[sys].get(tf, 0) for sys in sorted_systems
        )
        cell = ws.cell(row=total_row, column=tf_idx, value=col_total)
        cell.number_format = '$#,##0'
        cell.font = Font(bold=True)
        cell.border = thin_border

    grand_total = sum(
        cost
        for sys_costs in system_timeframe_costs.values()
        for cost in sys_costs.values()
    )
    gt_cell = ws.cell(row=total_row, column=len(headers), value=grand_total)
    gt_cell.number_format = '$#,##0'
    gt_cell.font = Font(bold=True, size=12)
    gt_cell.border = thin_border

    # Column widths
    ws.column_dimensions["A"].width = 28
    for col in range(2, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 18

    # Freeze header row
    ws.freeze_panes = "A2"

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    print(f"  Written: {output_path}")
